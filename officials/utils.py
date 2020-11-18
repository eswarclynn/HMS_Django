from institute.models import Block
from students.models import Attendance
from django.utils import timezone
from openpyxl import Workbook
from openpyxl import styles

class AttendanceBookGenerator:
    def __init__(self, block_id, year_month):
        self.block_id = block_id
        self.year_month = year_month
    
    def generate_workbook(self):
        workbook  = Workbook()
        workbook.remove(workbook.active)

        if self.block_id == 'all':
            for block in Block.objects.all():
                SheetGenerator = BlockAttendanceSheetGenerator(block.id, self.year_month)
                SheetGenerator.generate_block_sheet(workbook)
        else:
            SheetGenerator = BlockAttendanceSheetGenerator(self.block_id, self.year_month)
            SheetGenerator.generate_block_sheet(workbook)

        return workbook


class BlockAttendanceSheetGenerator:
    def __init__(self, block_id, year_month):
        self.block = Block.objects.get(id = block_id)
        self.attendance_list = Attendance.objects.filter(student__in = self.block.roomdetail_set.all().values_list('student', flat=True))
        if year_month == 'all':
            self.month = 'all'
            self.year = 'all'
        else:
            self.year, self.month = [int(x) for x in year_month.split("-")]

    def generate_block_sheet(self, workbook):
        worksheet = workbook.create_sheet(title = "{}".format(self.block.name))

        self.generate_dates()
        headers = ['Regd. No.', 'Name'] + [date.strftime("%d/%m/%y") for date in self.generated_dates]
        row_num = 1

        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold = True)
            cell.value = column_title

        for attendance in self.attendance_list:
            row_num += 1
            row_data = [attendance.student.regd_no, attendance.student.name] + self.get_student_attendance(attendance)

            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if cell_value == 'P': cell.font = styles.Font(color='28A745')
                elif cell_value == 'A': cell.font = styles.Font(bold=True, color='DC3545')


    def get_student_attendance(self, attendance):
        present_absent_list = []
        for date in self.generated_dates:
            date_formatted = date.strftime("%Y-%m-%d")
            if attendance.present_dates and attendance.present_dates.find(date_formatted) != -1:
                present_absent_list.append('P')
            elif attendance.absent_dates and attendance.absent_dates.find(date_formatted) != -1:
                present_absent_list.append('A')
            else:
                present_absent_list.append('-')

        return present_absent_list

    def generate_dates(self):
        if self.year == 'all' or self.month == 'all':
            date_set = self.get_marked_dates()
            month_set = self.get_month_year_set(date_set)
            generated_dates = []

            for item in month_set:
                dates_of_month = self.get_dates_of_month(item[0], item[1])
                generated_dates += dates_of_month

            self.generated_dates = sorted(generated_dates)
        
        else:
            self.generated_dates = self.get_dates_of_month(self.month, self.year)

    def get_dates_of_month(self, month, year):
            day = timezone.timedelta(days=1)
            start_date = timezone.datetime(year = year, month = month, day = 1)
            dates_of_month = []
            d = start_date
            while d.month == month:
                dates_of_month.append(d)
                d += day

            return sorted(dates_of_month)

    def get_marked_dates(self):
        date_set = set()
        for attendance in self.attendance_list:
            if attendance.present_dates:
                date_set |= set(attendance.present_dates.split(','))
            if attendance.absent_dates:
                date_set |= set(attendance.absent_dates.split(','))
        date_set.discard('')

        date_format = "%Y-%m-%d"
        date_set = set([timezone.datetime.strptime(date, date_format) for date in date_set])
        self.marked_dates = date_set
        return self.marked_dates

    # Set of tuples (month, year) that have attendance marked
    def get_month_year_set(self, date_set):
        month_year_set = set([(date.month, date.year) for date in date_set])
        return month_year_set